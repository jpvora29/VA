"""Render a :class:`mom.run_log.RunSummary` as a formatted two-sheet workbook.

Sheet 1 — Run Summary: the headline totals, then a per-phase table.
Sheet 2 — Call Log:    one row per LLM call.

Pure output: it reads a summary and writes a file, and knows nothing about the
pipeline that produced it. Falls back to JSON when openpyxl is unavailable.

Every string is scrubbed on its way into a cell. Call labels are built from slide
and section titles, and PowerPoint stores a soft line break inside a text frame as
a vertical tab (0x0b) -- which openpyxl refuses to write, and which has no visible
form to warn you with.
"""
from __future__ import annotations

import re
from pathlib import Path

from mom.run_log import RunSummary

# openpyxl's own rejection set (openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE), inlined so
# this module does not depend on a private import path that may move between releases.
ILLEGAL_IN_WORKSHEET = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")


def worksheet_safe(value):
    """A value openpyxl will accept: control characters out of strings, others as-is."""
    return ILLEGAL_IN_WORKSHEET.sub(" ", value) if isinstance(value, str) else value


NAVY = "002C77"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
TOTAL_GREY = "D9D9D9"
BORDER_GREY = "D9D9D9"

_SUMMARY_COLUMNS = (26, 14, 12, 14, 15, 18, 14)
_CALL_COLUMNS = (5, 22, 40, 16, 14, 15, 18, 13, 13)


def write_run_log(summary: RunSummary, path: Path) -> Path:
    """Write ``summary`` to ``path`` (.xlsx). Returns the path actually written."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
    except ImportError:
        fallback = path.with_suffix(".json")
        fallback.write_text(summary.model_dump_json(indent=2), encoding="utf-8")
        return fallback

    workbook = Workbook()
    _write_summary_sheet(workbook.active, summary)
    _write_call_sheet(workbook.create_sheet("Call Log"), summary)
    workbook.save(path)
    return path


# ── cell styling ──────────────────────────────────────────────────────────────


def _styles():
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color=BORDER_GREY)
    return {
        "header_font": Font(name="Arial", size=10, bold=True, color=WHITE),
        "header_fill": PatternFill("solid", fgColor=NAVY),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "alt_fill": PatternFill("solid", fgColor=LIGHT_GREY),
        "total_fill": PatternFill("solid", fgColor=TOTAL_GREY),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "Font": Font,
        "Alignment": Alignment,
    }


def _header(cell, text: str) -> None:
    s = _styles()
    cell.value = worksheet_safe(text)
    cell.font = s["header_font"]
    cell.fill = s["header_fill"]
    cell.alignment = s["header_align"]
    cell.border = s["border"]


def _cell(cell, value, *, bold=False, align="left", num_fmt=None, shade=False) -> None:
    s = _styles()
    cell.value = worksheet_safe(value)
    cell.font = s["Font"](name="Arial", size=10, bold=bold)
    cell.alignment = s["Alignment"](horizontal=align)
    cell.border = s["border"]
    if shade:
        cell.fill = s["alt_fill"]
    if num_fmt:
        cell.number_format = num_fmt


def _total_row(sheet, row: int, values: dict, columns: int) -> None:
    s = _styles()
    for col in range(1, columns + 1):
        cell = sheet.cell(row, col)
        cell.value = worksheet_safe(values.get(col))
        cell.font = s["Font"](name="Arial", size=10, bold=True)
        cell.fill = s["total_fill"]
        cell.border = s["border"]
        if isinstance(cell.value, int):
            cell.number_format = "#,##0"
            cell.alignment = s["Alignment"](horizontal="right")


def _widths(sheet, widths) -> None:
    from openpyxl.utils import get_column_letter

    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.row_dimensions[1].height = 30
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False


# ── sheets ────────────────────────────────────────────────────────────────────


def _write_summary_sheet(sheet, summary: RunSummary) -> None:
    sheet.title = "Run Summary"

    _header(sheet.cell(1, 1), "Metric")
    _header(sheet.cell(1, 2), "Value")
    headline = [
        ("Run ID", summary.run_id),
        ("Total pipeline time", f"{summary.total_duration_s}s"),
        ("Total LLM calls", summary.total_llm_calls),
        ("Total input tokens", summary.total_input_tokens),
        ("Total output tokens", summary.total_output_tokens),
        ("Total reasoning tokens", summary.total_reasoning_tokens),
        ("Total tokens", summary.total_tokens),
    ]
    row = 2
    for i, (label, value) in enumerate(headline):
        shade = i % 2 == 1
        _cell(sheet.cell(row, 1), label, bold=True, shade=shade)
        _cell(sheet.cell(row, 2), value, align="right", shade=shade,
              num_fmt="#,##0" if isinstance(value, int) else None)
        row += 1

    row += 1
    for col, head in enumerate(
        ["Phase", "Wall Time (s)", "LLM Calls", "Input Tokens",
         "Output Tokens", "Reasoning Tokens", "Total Tokens"], 1
    ):
        _header(sheet.cell(row, col), head)
    row += 1

    for i, phase in enumerate(summary.phase_breakdown):
        shade = i % 2 == 1
        values = [phase.phase, phase.wall_time_s, phase.calls, phase.input_tokens,
                  phase.output_tokens, phase.reasoning_tokens, phase.total_tokens]
        for col, value in enumerate(values, 1):
            numeric = isinstance(value, (int, float)) and col > 1
            _cell(
                sheet.cell(row, col), value,
                align="right" if numeric else "left",
                num_fmt=("0.00" if col == 2 else "#,##0" if isinstance(value, int) else None),
                shade=shade,
            )
        row += 1

    if summary.phase_breakdown:
        _total_row(sheet, row, {
            1: "TOTAL", 3: summary.total_llm_calls, 4: summary.total_input_tokens,
            5: summary.total_output_tokens, 6: summary.total_reasoning_tokens,
            7: summary.total_tokens,
        }, len(_SUMMARY_COLUMNS))

    _widths(sheet, _SUMMARY_COLUMNS)


def _write_call_sheet(sheet, summary: RunSummary) -> None:
    for col, head in enumerate(
        ["#", "Timestamp", "Label", "Phase", "Input Tokens", "Output Tokens",
         "Reasoning Tokens", "Total Tokens", "Duration (s)"], 1
    ):
        _header(sheet.cell(1, col), head)

    for i, call in enumerate(summary.calls):
        shade = i % 2 == 1
        values = [call.call_index, call.timestamp, call.label, call.phase,
                  call.usage.input_tokens, call.usage.output_tokens,
                  call.usage.reasoning_tokens, call.usage.total_tokens, call.duration_s]
        for col, value in enumerate(values, 1):
            is_int = isinstance(value, int) and col > 1
            is_float = isinstance(value, float)
            _cell(
                sheet.cell(i + 2, col), value,
                align="right" if (is_int or is_float) else "left",
                num_fmt="#,##0" if is_int else "0.00" if is_float else None,
                shade=shade,
            )

    if summary.calls:
        _total_row(sheet, len(summary.calls) + 2, {
            1: "TOTAL", 5: summary.total_input_tokens, 6: summary.total_output_tokens,
            7: summary.total_reasoning_tokens, 8: summary.total_tokens,
        }, len(_CALL_COLUMNS))

    _widths(sheet, _CALL_COLUMNS)
